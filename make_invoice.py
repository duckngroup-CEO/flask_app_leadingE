import pandas as pd
import openpyxl as xl
import shutil
import datetime
import app

# OUTPUT_FILEPATH = "static/img/output_files/output_data/output_data.xlsx"
# INVOICE_TEMPLATES  = "static/invoice_templates/invoice_template1.xlsx"
# OUTPUT_INVOICES_DIRPATH = "output_files/output_invoices"

# ファイルパスの指定
ORIGINAL_FILEPATH = "static/original_files/original_data.xlsx"
OUTPUT_FILEPATH = "static/output_files/output_data/output_data.xlsx"

INVOICE_TEMPLATES = "static/invoice_templates/invoice_template1.xlsx"
OUTPUT_INVOICES_DIRPATH = "static/output_files/output_invoices"

def main():
    # 請求書作成を実行
    make_invoice()

# エクセルを読み込む関数
def read_excel(filepath):
    df = pd.read_excel(filepath, index_col=0)
    return df

# データ上にある全ての顧客名（会社名）のリストを取得する（重複を消す）
def make_company_list():
    df = pd.read_excel(OUTPUT_FILEPATH)
    company_list = []
    company_list = list(df["顧客名"].unique())
    return company_list

# 各顧客別（会社別）の注文データを取得
def get_company_data(company):
    df=read_excel(OUTPUT_FILEPATH)
    df=df[df["顧客名"]==company]
    return df

# 請求書雛形を顧客の人数分にコピーして、名前を請求書名に変更
def copy_invoice_templates(company):
    shutil.copyfile(INVOICE_TEMPLATES, OUTPUT_INVOICES_DIRPATH + "/" + "【請求書】" + company + " 御中" +".xlsx")

# 顧客名と請求Noが入った辞書を作成関数{"あひる商事":"20220901-1"}
def make_number_dic(company_list):
    company_count = len(company_list)
    today = str(datetime.date.today().strftime('%Y%m%d'))

    # リスト生成[0,1,2,3,4・・・・]
    list = [x for x in range(1, company_count+1)] 

    # 辞書生成
    number_dic = {}
    for youso in list:
        number = today + "-" + str(list[youso-1])
        number_dic[company_list[youso-1]] = number
    return number_dic

# # リンク作成関数
# def make_linkname():
#     company_list = make_company_list()
#     linkname_list = []

#     for company in company_list:
#         link = f"static/output_files/output_invoices/【請求書】{company} 御中.xlsx"
#         linkname_list.append(link)
#     return linkname_list

# 請求書作成関数
def make_invoice():
    company_list = make_company_list()
    number_dic = make_number_dic(company_list)

    for company in company_list:
        df = get_company_data(company)
        copy_invoice_templates(company)
        wb = xl.load_workbook(OUTPUT_INVOICES_DIRPATH + "/" + "【請求書】" + company + " 御中" + ".xlsx")
        ws = wb.worksheets[0]

        # 会社名入れ込み
        ws["A2"].value = company

        # 請求日入れ込み
        ws["G3"].value = str(datetime.date.today())

        # ナンバー入れ込み （今日の日付-1）
        ws["G2"].value = number_dic[company]

        # 顧客ごとの商品一覧リスト作成
        each_df = df[df["顧客名"]==company]

        # 商品名と数量
        each_df  = each_df[["商品名", "数量","単価"]].groupby("商品名").agg(["sum","min"])

        each_data_list = []
        for data in each_df.itertuples():
            # print(list(data))
            each_data_list.append(list(data))
        # print(each_data_list)

        # 商品名、数量入れ込み
        if len(each_data_list)>14:
            print("データが14個以上なので、請求書に反映できません")
            pass
        else:
            for i in range(0, len(each_data_list)):
                # 商品名
                ws[f"A{i+15}"].value = each_data_list[i][0]
                # 数量
                ws[f"D{i+15}"].value = each_data_list[i][1]
                # 単価
                ws[f"F{i+15}"].value = each_data_list[i][4]
            # ブック保存
            filename = OUTPUT_INVOICES_DIRPATH + "/" + "【請求書】" + company + " 御中" + ".xlsx"
            wb.save(filename)
# メイン関数の実行
if __name__ == "__main__":
    main()
